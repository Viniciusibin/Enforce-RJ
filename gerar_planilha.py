"""
Gera a planilha padrão de atualização de ativos imobiliários.
Execute: python gerar_planilha.py
"""

import sqlite3
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── Paleta ──────────────────────────────────────────────────────────────────
AZUL_ESCURO  = "1F3864"
AZUL_MEDIO   = "2E75B6"
AZUL_CLARO   = "D9E2F3"
CINZA_CLARO  = "F2F2F2"
AMARELO      = "FFF2CC"
VERDE_CLARO  = "E2EFDA"
VERMELHO_CLR = "FCE4D6"
BRANCO       = "FFFFFF"


def estilo_cabecalho(ws, row, col, texto, cor_fundo=AZUL_ESCURO, cor_fonte=BRANCO, bold=True):
    cell = ws.cell(row=row, column=col, value=texto)
    cell.fill   = PatternFill("solid", fgColor=cor_fundo)
    cell.font   = Font(bold=bold, color=cor_fonte, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borda = Side(style="thin", color="AAAAAA")
    cell.border = Border(left=borda, right=borda, top=borda, bottom=borda)
    return cell


def estilo_dado(ws, row, col, valor=None, cor_fundo=BRANCO, formato=None):
    cell = ws.cell(row=row, column=col, value=valor)
    cell.fill = PatternFill("solid", fgColor=cor_fundo)
    cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    borda = Side(style="thin", color="D9D9D9")
    cell.border = Border(left=borda, right=borda, top=borda, bottom=borda)
    if formato:
        cell.number_format = formato
    return cell


def merge_header(ws, row, c_start, c_end, texto, cor_fundo=AZUL_MEDIO):
    ws.merge_cells(start_row=row, start_column=c_start, end_row=row, end_column=c_end)
    cell = ws.cell(row=row, column=c_start, value=texto)
    cell.fill = PatternFill("solid", fgColor=cor_fundo)
    cell.font = Font(bold=True, color=BRANCO, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell


# ── Definição das colunas ────────────────────────────────────────────────────
# (nome_exibição, campo_db, grupo, cor_grupo, largura, formato, dica)
COLUNAS = [
    # IDENTIFICAÇÃO
    ("ID / Matrícula",         "mat_tipo",             "IDENTIFICAÇÃO",   AZUL_CLARO,   22, None,          "Matrícula ou identificador único do imóvel"),
    ("Tipologia",              "tipologia",             "IDENTIFICAÇÃO",   AZUL_CLARO,   20, None,          "Ex.: Galpão, Fazenda, Apartamento..."),
    ("Cidade",                 "cidade",                "IDENTIFICAÇÃO",   AZUL_CLARO,   18, None,          "Município onde está o imóvel"),
    ("UF",                     "estado",                "IDENTIFICAÇÃO",   AZUL_CLARO,   6,  None,          "Sigla do estado (SP, RJ, MG...)"),
    # AVALIAÇÃO
    ("Valor de Mercado (VM) — Enforce", "vm_enforce",  "AVALIAÇÃO",       VERDE_CLARO,  22, 'R$ #,##0.00', "VM atual no sistema Enforce"),
    ("Valor de Venda Forçada (VF) — Enforce", "vf_enforce", "AVALIAÇÃO",  VERDE_CLARO,  28, 'R$ #,##0.00', "VF atual no sistema Enforce"),
    ("VF do Laudo",            "vf_laudo",              "AVALIAÇÃO",       VERDE_CLARO,  20, 'R$ #,##0.00', "VF conforme laudo de avaliação"),
    ("VF do Laudo Atualizado", "vf_laudo_att",          "AVALIAÇÃO",       VERDE_CLARO,  24, 'R$ #,##0.00', "VF atualizado após reavaliação"),
    ("Data do Laudo",          "dt_laudo",              "AVALIAÇÃO",       VERDE_CLARO,  16, "DD/MM/AAAA",  "Data de emissão do laudo"),
    # LAUDOS / DOCUMENTOS
    ("Laudo Lideratu",         "laudo_avaliacao_lideratu", "LAUDOS",      AMARELO,       24, None,          "Número/referência do laudo Lideratu"),
    ("Arquivo Lideratu",       "lideratu_arquivo",      "LAUDOS",          AMARELO,       24, None,          "Nome ou link do arquivo"),
    ("Laudo VM Sistema",       "laudo_vm_sistema",      "LAUDOS",          AMARELO,       24, None,          "Referência do laudo VM no sistema"),
    ("Laudo VF Sistema",       "laudo_vf_sistema",      "LAUDOS",          AMARELO,       24, None,          "Referência do laudo VF no sistema"),
    # STATUS
    ("Status Geral",           "status",                "STATUS",          CINZA_CLARO,  24, None,          "Situação atual do ativo"),
    ("Obs. Status",            "obs_status",            "STATUS",          CINZA_CLARO,  30, None,          "Detalhes sobre o status"),
    ("Status Tarefa",          "status_tarefa",         "STATUS",          CINZA_CLARO,  20, None,          "Status da tarefa em aberto"),
    ("Data Criação Tarefa",    "dt_criacao_tarefa",     "STATUS",          CINZA_CLARO,  20, "DD/MM/AAAA",  "Data em que a tarefa foi criada"),
    # COMERCIAL
    ("Proposta (R$)",          "proposta",              "COMERCIAL",       VERMELHO_CLR, 20, 'R$ #,##0.00', "Valor da proposta recebida"),
    ("Comprador",              "comprador",             "COMERCIAL",       VERMELHO_CLR, 24, None,          "Nome do comprador ou interessado"),
    ("Tabela GI",              "tabela_gi",             "COMERCIAL",       VERMELHO_CLR, 16, None,          "Referência na tabela GI"),
    ("Valor Garantia (R$)",    "valor_garantia",        "COMERCIAL",       VERMELHO_CLR, 20, 'R$ #,##0.00', "Valor da garantia vinculada"),
    ("Natureza",               "natureza",              "COMERCIAL",       VERMELHO_CLR, 18, None,          "Natureza do negócio"),
    # VENDA
    ("Data da Venda",          "data_venda",            "VENDA",           AZUL_CLARO,   16, "DD/MM/AAAA",  "Data de fechamento da venda"),
    ("Valor Total da Venda (R$)", "valor_venda_total",  "VENDA",           AZUL_CLARO,   24, 'R$ #,##0.00', "Valor total recebido na venda"),
    ("Fluxo da Venda",         "fluxo_venda",           "VENDA",           AZUL_CLARO,   24, None,          "Descrição do fluxo financeiro"),
    ("Pendência Venda",        "pendencia_venda",       "VENDA",           AZUL_CLARO,   28, None,          "Pendências para conclusão da venda"),
    ("Responsável Pendência",  "quem_pendencia_venda",  "VENDA",           AZUL_CLARO,   22, None,          "Quem deve resolver a pendência"),
    ("Previsão de Registro",   "previsao_registro",     "VENDA",           AZUL_CLARO,   20, "DD/MM/AAAA",  "Data prevista para registro em cartório"),
    # OBSERVAÇÕES
    ("Observações Gerais",     "obs",                   "OBS",             CINZA_CLARO,  40, None,          "Qualquer informação adicional relevante"),
]


def cor_grupo(grupo):
    mapa = {
        "IDENTIFICAÇÃO": AZUL_CLARO,
        "AVALIAÇÃO":     VERDE_CLARO,
        "LAUDOS":        AMARELO,
        "STATUS":        CINZA_CLARO,
        "COMERCIAL":     VERMELHO_CLR,
        "VENDA":         AZUL_CLARO,
        "OBS":           CINZA_CLARO,
    }
    return mapa.get(grupo, BRANCO)


def criar_aba_dados(wb, imoveis_db):
    ws = wb.active
    ws.title = "Atualização de Ativos"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "E4"  # congela identificação e rola o resto

    # ── Linha 1: título principal ────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUNAS))
    titulo = ws.cell(row=1, column=1, value="ENFORCE — Atualização de Ativos Imobiliários")
    titulo.fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    titulo.font = Font(bold=True, color=BRANCO, size=14)
    titulo.alignment = Alignment(horizontal="center", vertical="center")

    # ── Linha 2: grupos de colunas ───────────────────────────────────────────
    ws.row_dimensions[2].height = 20

    # Calcular faixas de grupos
    grupo_faixas = {}
    for i, (_, _, grupo, *_) in enumerate(COLUNAS, start=1):
        if grupo not in grupo_faixas:
            grupo_faixas[grupo] = [i, i]
        else:
            grupo_faixas[grupo][1] = i

    for grupo, (c_ini, c_fim) in grupo_faixas.items():
        merge_header(ws, 2, c_ini, c_fim, grupo)

    # ── Linha 3: cabeçalhos das colunas ─────────────────────────────────────
    ws.row_dimensions[3].height = 40
    for i, (nome, campo, grupo, cor, largura, fmt, dica) in enumerate(COLUNAS, start=1):
        estilo_cabecalho(ws, 3, i, nome, cor_fundo=AZUL_ESCURO if i <= 4 else AZUL_MEDIO)
        ws.column_dimensions[get_column_letter(i)].width = largura

    # ── Validações de dados ──────────────────────────────────────────────────
    # Coluna "status" → lista
    col_status = next(i for i,(n,c,*_) in enumerate(COLUNAS,1) if c == "status")
    dv_status = DataValidation(
        type="list",
        formula1='"Não Vendido,Vendido,Proposta em Negociação,Em Avaliação,Cancelado"',
        allow_blank=True, showErrorMessage=True,
        errorTitle="Valor inválido", error="Use a lista suspensa para selecionar o status."
    )
    ws.add_data_validation(dv_status)
    dv_status.sqref = f"{get_column_letter(col_status)}4:{get_column_letter(col_status)}500"

    # ── Dados do banco ───────────────────────────────────────────────────────
    campos_db = [c[1] for c in COLUNAS]
    cores_linha = [BRANCO, CINZA_CLARO]
    for r_idx, row in enumerate(imoveis_db, start=4):
        cor = cores_linha[r_idx % 2]
        ws.row_dimensions[r_idx].height = 18
        for c_idx, campo in enumerate(campos_db, start=1):
            val = row.get(campo)
            fmt = COLUNAS[c_idx - 1][5]
            estilo_dado(ws, r_idx, c_idx, val, cor_fundo=cor, formato=fmt)

    # Linhas em branco para preenchimento (caso planilha enviada sem dados)
    linhas_preenchidas = len(imoveis_db)
    for r_idx in range(4 + linhas_preenchidas, 4 + linhas_preenchidas + 20):
        cor = cores_linha[r_idx % 2]
        ws.row_dimensions[r_idx].height = 18
        for c_idx in range(1, len(COLUNAS) + 1):
            fmt = COLUNAS[c_idx - 1][5]
            estilo_dado(ws, r_idx, c_idx, cor_fundo=cor, formato=fmt)

    # ── Rodapé / instrução rápida ────────────────────────────────────────────
    ultima_linha = 4 + linhas_preenchidas + 21
    ws.merge_cells(start_row=ultima_linha, start_column=1, end_row=ultima_linha, end_column=len(COLUNAS))
    rodape = ws.cell(row=ultima_linha, column=1,
                     value="⚠  Preencha somente as colunas que precisam de atualização. "
                           "Não altere o ID / Matrícula. Dúvidas: ver aba 'Instruções'.")
    rodape.fill = PatternFill("solid", fgColor=AMARELO)
    rodape.font = Font(bold=True, color="7F6000", size=10, italic=True)
    rodape.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[ultima_linha].height = 22


def criar_aba_instrucoes(wb):
    ws = wb.create_sheet("Instruções")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70

    # Título
    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = "INSTRUÇÕES DE PREENCHIMENTO"
    t.fill  = PatternFill("solid", fgColor=AZUL_ESCURO)
    t.font  = Font(bold=True, color=BRANCO, size=14)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    instrucoes = [
        ("REGRA GERAL", ""),
        ("ID / Matrícula", "NÃO altere este campo — é a chave de identificação do imóvel no sistema."),
        ("Campos em branco", "Deixe em branco os campos que não precisam de atualização. Não apague valores existentes sem certeza."),
        ("Datas", "Use o formato DD/MM/AAAA (ex.: 15/03/2026)."),
        ("Valores monetários", "Números apenas, sem R$ ou pontos de milhar (ex.: 1500000.00)."),
        ("", ""),
        ("STATUS GERAL — valores aceitos", ""),
        ("Não Vendido",                "Imóvel ainda não foi alienado."),
        ("Vendido",                    "Venda concluída e registrada."),
        ("Proposta em Negociação",     "Há proposta formal em análise."),
        ("Em Avaliação",               "Aguardando laudo ou reavaliação."),
        ("Cancelado",                  "Processo de venda encerrado sem êxito."),
        ("", ""),
        ("CAMPOS DE AVALIAÇÃO", ""),
        ("VM Enforce",                 "Valor de mercado registrado na plataforma Enforce."),
        ("VF Enforce",                 "Valor de venda forçada registrado na plataforma Enforce."),
        ("VF do Laudo",                "Valor de venda forçada conforme laudo técnico assinado."),
        ("VF do Laudo Atualizado",     "VF revisado após atualização monetária ou nova avaliação."),
        ("Data do Laudo",              "Data de assinatura do laudo pelo avaliador."),
        ("", ""),
        ("COMERCIAL", ""),
        ("Proposta (R$)",              "Valor da proposta recebida do interessado."),
        ("Comprador",                  "Nome completo ou razão social do comprador/interessado."),
        ("Valor Garantia",             "Valor do bem dado em garantia vinculado ao imóvel."),
        ("", ""),
        ("VENDA", ""),
        ("Data da Venda",              "Data de assinatura da escritura ou contrato definitivo."),
        ("Valor Total da Venda",       "Valor total efetivamente recebido ou acordado."),
        ("Fluxo da Venda",             "Descreva o parcelamento ou forma de pagamento."),
        ("Pendência Venda",            "Liste pendências que impedem ou retardam o registro."),
        ("Responsável Pendência",      "Nome do responsável por resolver a pendência."),
        ("Previsão de Registro",       "Data prevista para protocolo no cartório de imóveis."),
        ("", ""),
        ("DÚVIDAS?", "Entre em contato com o gestor de portfólio antes de enviar a planilha."),
    ]

    borda = Side(style="thin", color="DDDDDD")
    for r, (campo, desc) in enumerate(instrucoes, start=2):
        ws.row_dimensions[r].height = 18
        c1 = ws.cell(row=r, column=1, value=campo)
        c2 = ws.cell(row=r, column=2, value=desc)

        if campo in ("REGRA GERAL", "STATUS GERAL — valores aceitos",
                     "CAMPOS DE AVALIAÇÃO", "COMERCIAL", "VENDA", "DÚVIDAS?"):
            for c in (c1, c2):
                c.fill = PatternFill("solid", fgColor=AZUL_MEDIO)
                c.font = Font(bold=True, color=BRANCO, size=10)
        else:
            c1.font = Font(bold=True, size=10)
            c2.font = Font(size=10)
            bg = CINZA_CLARO if r % 2 == 0 else BRANCO
            for c in (c1, c2):
                c.fill = PatternFill("solid", fgColor=bg)

        for c in (c1, c2):
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = Border(left=borda, right=borda, top=borda, bottom=borda)


def carregar_imoveis():
    try:
        conn = sqlite3.connect("instance/prj.db")
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM imoveis ORDER BY id")
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        print(f"  {len(rows)} imóveis carregados do banco.")
        return rows
    except Exception as e:
        print(f"  Aviso: banco não encontrado ({e}). Planilha gerada sem dados pré-preenchidos.")
        return []


def main():
    print("Gerando planilha de atualização de ativos imobiliários...")
    imoveis = carregar_imoveis()

    wb = Workbook()
    criar_aba_dados(wb, imoveis)
    criar_aba_instrucoes(wb)

    nome_arquivo = "Atualizacao_Ativos_Imobiliarios.xlsx"
    wb.save(nome_arquivo)
    print(f"  Planilha salva: {nome_arquivo}")
    print("Pronto!")


if __name__ == "__main__":
    main()
