from __future__ import annotations

import threading
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COZAC_PATH = Path(__file__).parent / "dados cozac.xlsx"
COTIA_PATH  = Path(__file__).parent / "dados_cotia.xlsx"
# IDs de Cotia são retornados como (real_id + COTIA_OFFSET) para nunca colidirem com Cozac
COTIA_OFFSET = 10_000

_lock = threading.Lock()

IMOVEIS_COLS = [
    "id", "mat_tipo", "tipologia", "cidade", "estado",
    "vf_tombamento", "vm_tombamento", "resultado_localizacao",
    "area_terreno", "unid_area_terreno", "cidade_imovel",
    "classe_imovel", "tipo_imovel",
    "valor_avaliacao", "valor_venda_forcada", "valor_venda_forcada_vp",
    "laudo_avaliacao_lideratu", "lideratu_arquivo", "laudo_vm_sistema", "laudo_vf_sistema",
    "nivel", "vm_enforce", "vf_enforce", "vf_laudo", "dt_laudo", "obs",
    "status_tarefa", "dt_criacao_tarefa", "vf_laudo_att",
    "status", "obs_status",
    "tabela_gi", "proposta", "natureza", "valor_garantia", "comprador", "previsao_registro",
    "lat", "lng",
    "data_venda", "valor_venda_total", "fluxo_venda", "pendencia_venda", "quem_pendencia_venda",
]

PROPOSTAS_COLS = [
    "id", "matricula", "valor_venda", "fluxo", "quem_fez",
    "pendencia", "quem_pendencia", "obs", "created_at",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_COL_WIDTHS = {
    "id": 6, "mat_tipo": 18, "tipologia": 15, "cidade": 16, "estado": 8,
    "vf_tombamento": 14, "vm_tombamento": 14, "resultado_localizacao": 20,
    "area_terreno": 14, "unid_area_terreno": 12, "cidade_imovel": 16,
    "classe_imovel": 14, "tipo_imovel": 14,
    "valor_avaliacao": 16, "valor_venda_forcada": 18, "valor_venda_forcada_vp": 20,
    "laudo_avaliacao_lideratu": 22, "lideratu_arquivo": 18,
    "laudo_vm_sistema": 18, "laudo_vf_sistema": 18,
    "nivel": 10, "vm_enforce": 12, "vf_enforce": 12, "vf_laudo": 12,
    "dt_laudo": 12, "obs": 30,
    "status_tarefa": 16, "dt_criacao_tarefa": 16, "vf_laudo_att": 14,
    "status": 14, "obs_status": 24,
    "tabela_gi": 12, "proposta": 14, "natureza": 14, "valor_garantia": 14,
    "comprador": 20, "previsao_registro": 18,
    "lat": 12, "lng": 12,
    "data_venda": 12, "valor_venda_total": 16,
    "fluxo_venda": 14, "pendencia_venda": 20, "quem_pendencia_venda": 20,
}


def _style_header_row(ws, cols: list[str]) -> None:
    for i, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(i)].width = _COL_WIDTHS.get(col, 14)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _create_file(path: Path) -> None:
    wb = Workbook()
    ws_i = wb.active
    ws_i.title = "Imoveis"
    _style_header_row(ws_i, IMOVEIS_COLS)
    ws_p = wb.create_sheet("Propostas")
    _style_header_row(ws_p, PROPOSTAS_COLS)
    wb.save(path)


def _load_sheet(path: Path, sheet_name: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = [str(c) if c is not None else "" for c in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append(dict(zip(header, row)))
    return result


def _save_sheet(path: Path, sheet_name: str, cols: list[str], data_rows: list[dict]) -> None:
    wb = load_workbook(path)
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    for row_idx, row in enumerate(data_rows, start=2):
        for col_idx, col in enumerate(cols, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col))
    wb.save(path)


def _file_and_real_id(logical_id: int) -> tuple[Path, int]:
    """Devolve (arquivo, id_real_no_arquivo) a partir do id lógico."""
    if logical_id >= COTIA_OFFSET:
        return COTIA_PATH, logical_id - COTIA_OFFSET
    return COZAC_PATH, logical_id


def init_db() -> None:
    if not COZAC_PATH.exists():
        _create_file(COZAC_PATH)
    if not COTIA_PATH.exists():
        _create_file(COTIA_PATH)


# ---------------------------------------------------------------------------
# Imoveis
# ---------------------------------------------------------------------------

def get_all_imoveis() -> list[dict]:
    init_db()
    result = []
    for path, offset in [(COZAC_PATH, 0), (COTIA_PATH, COTIA_OFFSET)]:
        if not path.exists():
            continue
        for row in _load_sheet(path, "Imoveis"):
            r = dict(row)
            if r.get("id") is not None:
                r["id"] = int(r["id"]) + offset
            result.append(r)
    return result


def get_imovel(logical_id: int) -> dict | None:
    path, real_id = _file_and_real_id(logical_id)
    for r in _load_sheet(path, "Imoveis"):
        if r.get("id") == real_id:
            row = dict(r)
            row["id"] = logical_id
            return row
    return None


def get_imovel_by_mat(mat: str) -> dict | None:
    mat = str(mat).strip()
    for path, offset in [(COZAC_PATH, 0), (COTIA_PATH, COTIA_OFFSET)]:
        if not path.exists():
            continue
        for r in _load_sheet(path, "Imoveis"):
            if str(r.get("mat_tipo") or "").strip() == mat:
                row = dict(r)
                if row.get("id") is not None:
                    row["id"] = int(row["id"]) + offset
                return row
    return None


def insert_imovel(data: dict) -> int:
    # Novos imóveis vão para Cozac por padrão
    with _lock:
        rows = _load_sheet(COZAC_PATH, "Imoveis")
        new_id = max((int(r.get("id") or 0) for r in rows), default=0) + 1
        entry = dict(data)
        entry["id"] = new_id
        rows.append(entry)
        _save_sheet(COZAC_PATH, "Imoveis", IMOVEIS_COLS, rows)
    return new_id


def update_imovel(logical_id: int, data: dict) -> bool:
    path, real_id = _file_and_real_id(logical_id)
    with _lock:
        rows = _load_sheet(path, "Imoveis")
        updated = False
        for r in rows:
            if r.get("id") == real_id:
                r.update({k: v for k, v in data.items() if k != "id"})
                updated = True
                break
        if updated:
            _save_sheet(path, "Imoveis", IMOVEIS_COLS, rows)
    return updated


def delete_imovel(logical_id: int) -> bool:
    path, real_id = _file_and_real_id(logical_id)
    with _lock:
        rows = _load_sheet(path, "Imoveis")
        new_rows = [r for r in rows if r.get("id") != real_id]
        if len(new_rows) == len(rows):
            return False
        _save_sheet(path, "Imoveis", IMOVEIS_COLS, new_rows)
    return True


# ---------------------------------------------------------------------------
# Propostas (ficam em Cozac)
# ---------------------------------------------------------------------------

def get_all_propostas() -> list[dict]:
    init_db()
    return _load_sheet(COZAC_PATH, "Propostas")


def get_proposta(proposta_id: int) -> dict | None:
    for r in get_all_propostas():
        if r.get("id") == proposta_id:
            return r
    return None


def insert_proposta(data: dict) -> int:
    with _lock:
        rows = get_all_propostas()
        new_id = max((int(r.get("id") or 0) for r in rows), default=0) + 1
        entry = dict(data)
        entry["id"] = new_id
        if "created_at" not in entry:
            entry["created_at"] = datetime.now().isoformat(timespec="seconds")
        rows.append(entry)
        _save_sheet(COZAC_PATH, "Propostas", PROPOSTAS_COLS, rows)
    return new_id


def update_proposta(proposta_id: int, data: dict) -> bool:
    allowed = {"matricula", "valor_venda", "fluxo", "quem_fez", "pendencia", "quem_pendencia", "obs"}
    with _lock:
        rows = get_all_propostas()
        updated = False
        for r in rows:
            if r.get("id") == proposta_id:
                r.update({k: v for k, v in data.items() if k in allowed})
                updated = True
                break
        if updated:
            _save_sheet(COZAC_PATH, "Propostas", PROPOSTAS_COLS, rows)
    return updated


if __name__ == "__main__":
    init_db()
    print(f"Cozac: {COZAC_PATH}")
    print(f"Cotia: {COTIA_PATH}")
